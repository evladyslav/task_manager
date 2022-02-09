from openpyxl import Workbook, load_workbook
from datetime import datetime
import telebot
import os
from config import token, PATH


if os.path.isfile(PATH):
    workbook = load_workbook(filename=PATH, data_only=True)
    print("File exists and will be overwritten NOW")
else:
    print("The file is missing, new one is created")
    workbook = Workbook()

oil = telebot.TeleBot(token)
schedule = {}
start = 0
msg_id = 0


@oil.message_handler(commands=['file'])
def send_file(message):
    username = '{0}'.format(str(message.from_user.first_name))

    if str(message.from_user.first_name) in workbook.sheetnames:
        def copier():
            wb = Workbook()
            wb.create_sheet(username)
            mr = workbook[username].max_row
            mc = workbook[username].max_column
            for i in range(1, mr + 1):
                for j in range(1, mc + 1):
                    c = workbook[username].cell(row=i, column=j)
                    wb[username].cell(row=i, column=j).value = c.value
            wb.save('{0}.xlsx'.format(username))

        if message.from_user.id == 608635889:
            with open(PATH, 'rb') as file:
                oil.send_document(message.chat.id, file)
        elif message.from_user.id == 67968372:
            with open(PATH, 'rb') as file:
                oil.send_document(message.chat.id, file)
        else:
            copier()
            with open('{0}.xlsx'.format(username), "rb") as xlsxfile:
                oil.send_document(message.chat.id, xlsxfile)
            os.remove('{0}.xlsx'.format(username))
    else:
        oil.send_message(message.chat.id, text='File does not exist.. ')


@oil.message_handler(commands=['help'])
def help_message(message):
    help_text = '1. To use this bot send name of the task then press button "Start"\n' \
        '2. /file to get .xlsx file of all tasks'
    oil.send_message(message.chat.id, help_text)


@oil.message_handler(commands=['start'])
def start_message(message):
    start_text = 'Welcome! This bot is used to record data about your activity during the day.' \
        '\nUse /help for more info '
    oil.send_message(message.chat.id, start_text)


@oil.message_handler(content_types=['text'])
def analyzer(message):
    global msg_id
    user = str(message.from_user.first_name) + ' ' + str(message.from_user.last_name)
    print(user, ': connected')
    msg_id = message.message_id
    print(message.from_user.id, user)
    schedule[msg_id + 1] = []
    schedule[msg_id + 1].append(user)
    date_stamp = str(datetime.date(datetime.today()))
    schedule[msg_id + 1].append(date_stamp)
    task = str(message.text)
    schedule[msg_id + 1].append(task)
    markup = telebot.types.InlineKeyboardMarkup(row_width=2)
    markup.row(telebot.types.InlineKeyboardButton(text='Start', callback_data="Begin"))
    # oil_photo = open('ooo.jpg', "rb")
    # oil.send_photo(message.chat.id, oil_photo)
    oil.send_message(message.chat.id, text=message.text, reply_markup=markup)


def put_in(fin, tme, msgid):
    schedule[msgid].append(('Конец: ' + str(datetime.time(fin))))
    schedule[msgid].append(str(tme))


@oil.callback_query_handler(func=lambda call: True)
def query_handler(call):
    global start
    msg_txt = call.message.text

    def duration(beg, fin):
        dur = fin - beg
        return dur

    if call.data == 'Begin':
        oil.answer_callback_query(callback_query_id=call.id, text='Started!')
        key_finish = telebot.types.InlineKeyboardMarkup(row_width=2)
        key_finish.row(telebot.types.InlineKeyboardButton(text='Finish', callback_data="Finish"))

        start = datetime.now()
        oil.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text=call.message.text +
                              '\nStart time: {0}'.format(str(start)),
                              reply_markup=key_finish)
        schedule[call.message.message_id].append(('Начало: ' + str(datetime.time(start))))
        print(call.from_user.first_name, ' started task')
    elif call.data == 'Finish':
        end = datetime.now()
        time = duration(start, end)

        oil.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                              text=msg_txt + '\nFinish: {0}'.format(str(end)) + '\nWasted time: {0}'.format(time))
        print(call.from_user.first_name, ' finished task')
        put_in(end, time, call.message.message_id)

        if str(call.from_user.first_name) not in workbook.sheetnames:
            workbook.create_sheet(title='{0}'.format(str(call.from_user.first_name)))
            temp_sheet = workbook['{0}'.format(str(call.from_user.first_name))]
            temp_sheet.append(schedule[call.message.message_id])

        else:
            temp_sheet = workbook['{0}'.format(str(call.from_user.first_name))]
            temp_sheet.append(schedule[call.message.message_id])

        wasted = 'Wasted time: {0}'.format(time)
        workbook.save(filename=PATH)
        oil.answer_callback_query(callback_query_id=call.id, text=wasted)


if __name__ == '__main__':
    oil.polling(none_stop=True)
